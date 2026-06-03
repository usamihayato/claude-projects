# =============================================================
# Snowflake Notebook: Excel ワークスペースアップ → テーブル保存
# =============================================================
#
# 【実行手順】
#   1. ワークスペースに Excel ファイルをアップロード
#   2. Cell 3 の FILE_NAME を実際のファイル名に変更
#   3. 初回のみ Cell 1〜Cell 3 をすべて実行
#      2回目以降は Cell 1 と Cell 3 だけ実行すれば OK
#
# 【テーブル構成】
#   EXCEL_UPLOADS
#     FILE_NAME    VARCHAR   : アップロードしたファイル名
#     SHEET_NAME   VARCHAR   : シート名
#     UPLOADED_AT  TIMESTAMP : 保存日時（自動設定）
#     CONTENT      TEXT      : 当該シートのパース済みテキスト（行データ）
#     ROW_COUNT    NUMBER    : 有効行数
# =============================================================


# ---- Cell 1: ライブラリ インポート & パース関数定義 ----
# ※ ここに既存の extract_from_excel 関数をそのまま貼り付けてください

import io
from datetime import datetime
import openpyxl


def extract_from_excel(buffer: bytes, include_sheet_names: bool = True) -> str:
    """
    既存のパース関数をここに貼り付けてください。
    シグネチャ: (buffer: bytes, include_sheet_names: bool) -> str
    戻り値   : シート名＋行データを改行で結合した文字列
    """
    # ▼▼▼ 既存コードをここに貼り付け ▼▼▼
    wb = openpyxl.load_workbook(io.BytesIO(buffer), data_only=True)
    all_lines = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if include_sheet_names:
            all_lines.append(f"[Sheet: {sheet_name}]")
        for row in ws.iter_rows(values_only=True):
            line = "\t".join("" if v is None else str(v) for v in row)
            all_lines.append(line)

    return "\n".join(all_lines)
    # ▲▲▲ 既存コードここまで ▲▲▲


def extract_sheets_from_excel(buffer: bytes) -> dict[str, str]:
    """シートごとのテキストを {シート名: 行データ文字列} の辞書で返す"""
    wb = openpyxl.load_workbook(io.BytesIO(buffer), data_only=True)
    sheets: dict[str, str] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines = []
        for row in ws.iter_rows(values_only=True):
            line = "\t".join("" if v is None else str(v) for v in row)
            lines.append(line)
        sheets[sheet_name] = "\n".join(lines)

    return sheets


# ---- Cell 2: テーブル作成（初回のみ実行） ----

session.sql("""
    CREATE TABLE IF NOT EXISTS EXCEL_UPLOADS (
        FILE_NAME    VARCHAR(512)  NOT NULL,
        SHEET_NAME   VARCHAR(512)  NOT NULL,
        UPLOADED_AT  TIMESTAMP_NTZ DEFAULT CURRENT_TIMESTAMP(),
        CONTENT      TEXT,
        ROW_COUNT    NUMBER(10, 0)
    )
""").collect()

print("テーブル準備完了")


# ---- Cell 3: メインロジック ----

# ▼ アップロードしたファイル名を指定
FILE_NAME = "sample.xlsx"

# ワークスペースにアップしたファイルは /workspace/ にマウントされる
file_path = f"/workspace/{FILE_NAME}"

with open(file_path, "rb") as f:
    buffer = f.read()

# シートごとにパース
sheets = extract_sheets_from_excel(buffer)

# シートごとにレコードを作成
from snowflake.snowpark import Row

uploaded_at = datetime.now()
rows = [
    Row(
        FILE_NAME=FILE_NAME,
        SHEET_NAME=sheet_name,
        UPLOADED_AT=uploaded_at,
        CONTENT=content,
        ROW_COUNT=sum(1 for line in content.split("\n") if line.strip()),
    )
    for sheet_name, content in sheets.items()
]

df = session.create_dataframe(rows)
df.write.mode("append").save_as_table("EXCEL_UPLOADS")

print(f"保存完了: {FILE_NAME}")
for sheet_name, content in sheets.items():
    row_count = sum(1 for line in content.split("\n") if line.strip())
    print(f"  - {sheet_name}: {row_count:,} 行")
